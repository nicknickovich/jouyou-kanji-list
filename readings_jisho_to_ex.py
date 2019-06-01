import openpyxl
from readings_jisho import kanji_dict
from kanji_all import kanji_list

wb = openpyxl.Workbook()
sheet = wb["Sheet"]

sheet["A1"] = "Kanji"
sheet["B1"] = "On"
sheet["C1"] = "Kun"
sheet["D1"] = "Meaning"

for i in range(2, len(kanji_list) + 2):
    item = kanji_list[i-2]
    sheet.cell(row=i, column=1).value = item
    sheet.cell(row=i, column=2).value = ", ".join(kanji_dict[item]["on"])
    sheet.cell(row=i, column=3).value = ", ".join(kanji_dict[item]["kun"])
    sheet.cell(row=i, column=4).value = ", ".join(kanji_dict[item]["meanings"])

# make first row alvays visible
sheet.freeze_panes = "A2"
wb.save("readings_jisho.xlsx")